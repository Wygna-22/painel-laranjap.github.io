from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def gerar_excel_colaboradores(colaboradores):
    """
    Gera um relatório Excel com os colaboradores.
    """
    # Cria a planilha
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Colaboradores"

    # ==========================
    # TÍTULO
    # ==========================
    worksheet.merge_cells("A1:J1")

    titulo = worksheet["A1"]
    titulo.value = "RELATÓRIO DE COLABORADORES"

    titulo.font = Font(
        bold=True,
        size=16,
        color="FFFFFF"
    )

    titulo.fill = PatternFill(
        fill_type="solid",
        fgColor="F28C28"
    )

    titulo.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    # ==========================
    # CABEÇALHOS
    # ==========================
    cabecalhos = [
        "Nome",
        "Matrícula",
        "E-mail",
        "Telefone",
        "Cidade",
        "Setor",
        "Cargo",
        "Gestor",
        "Status",
        "Admissão",
    ]

    for coluna, texto in enumerate(cabecalhos, start=1):

        celula = worksheet.cell(
            row=3,
            column=coluna
        )

        celula.value = texto

        celula.font = Font(
            bold=True,
            color="FFFFFF"
        )

        celula.fill = PatternFill(
            fill_type="solid",
            fgColor="F28C28"
        )

        celula.alignment = Alignment(
            horizontal="center"
        )

    # ==========================
    # DADOS
    # ==========================
    linha = 4

    for colaborador in colaboradores:

        worksheet.cell(row=linha, column=1).value = colaborador.nome
        worksheet.cell(row=linha, column=2).value = colaborador.matricula
        worksheet.cell(row=linha, column=3).value = colaborador.email
        worksheet.cell(row=linha, column=4).value = colaborador.telefone
        worksheet.cell(row=linha, column=5).value = colaborador.cidade
        worksheet.cell(row=linha, column=6).value = colaborador.setor
        worksheet.cell(row=linha, column=7).value = colaborador.cargo

        # Caso o colaborador possua gestor relacionado
        worksheet.cell(
            row=linha,
            column=8
        ).value = (
            colaborador.gestor.nome
            if colaborador.gestor
            else ""
        )

        worksheet.cell(
            row=linha,
            column=9
        ).value = colaborador.status.value

        worksheet.cell(
            row=linha,
            column=10
        ).value = colaborador.data_admissao.strftime("%d/%m/%Y")

        linha += 1

    # ==========================
    # AJUSTAR LARGURA DAS COLUNAS
    # ==========================
    larguras = {
        "A": 30,
        "B": 15,
        "C": 35,
        "D": 18,
        "E": 20,
        "F": 20,
        "G": 20,
        "H": 25,
        "I": 18,
        "J": 18,
    }

    for coluna, largura in larguras.items():
        worksheet.column_dimensions[coluna].width = largura

    # ==========================
    # SALVAR EM MEMÓRIA
    # ==========================

    arquivo = BytesIO()

    workbook.save(arquivo)

    arquivo.seek(0)

    return arquivo