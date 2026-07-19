from abc import ABC, abstractmethod
from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment,
)
from openpyxl.utils import get_column_letter


class BaseExcelReport(ABC):

    def generate(self) -> BytesIO:
        workbook = Workbook()
        worksheet = workbook.active

        worksheet.title = self.get_title()

        self._write_title(worksheet)
        self._write_statistics(worksheet)
        self._write_headers(worksheet)
        self._write_rows(worksheet)
        self._format_columns(worksheet)
        self._write_footer(worksheet)

        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)

        return stream

    @abstractmethod
    def get_title(self):
        ...

    @abstractmethod
    def get_headers(self):
        ...

    @abstractmethod
    def get_rows(self):
        ...

    @abstractmethod
    def get_footer(self):
        ...

    def get_statistics(self) -> list[tuple[str, str]]:
        return []

    def _write_title(self, ws):

        total_columns = len(self.get_headers())

        ws.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=total_columns,
        )

        title = ws["A1"]

        title.value = self.get_title()

        title.font = Font(
            bold=True,
            size=18,
            color="FFFFFF",
        )

        title.fill = PatternFill(
            fill_type="solid",
            fgColor="F97316",
        )

        title.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        ws.row_dimensions[1].height = 30

        ws["A2"] = (
            f"Gerado em: "
            f"{datetime.now():%d/%m/%Y %H:%M}"
        )

        ws["A2"].font = Font(
            italic=True,
            color="666666",
        )

    def _write_statistics(self, ws):

        statistics = self.get_statistics()

        if not statistics:
            return

        row = 3

        for label, value in statistics:

            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=1).font = Font(bold=True)

            ws.cell(row=row, column=2).value = value

            row += 1

    def _write_headers(self, ws):

        headers = self.get_headers()

        row = 8

        fill = PatternFill(
            fill_type="solid",
            fgColor="EA580C",
        )

        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col, value in enumerate(headers, start=1):

            cell = ws.cell(row=row, column=col)

            cell.value = value
            cell.font = Font(
                bold=True,
                color="FFFFFF",
            )

            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

    def _write_rows(self, ws):

        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        fill = PatternFill(
            fill_type="solid",
            fgColor="F8FAFC",
        )

        rows = self.get_rows()

        start = 9

        for index, row_data in enumerate(rows):

            row_number = start + index
            ws.row_dimensions[row_number].height = 22
            
            for column, value in enumerate(row_data, start=1):

                cell = ws.cell(
                    row=row_number,
                    column=column,
                )

                cell.value = value
                cell.border = border
                cell.alignment = Alignment(
                    vertical="center",
                )

                if index % 2 == 0:
                    cell.fill = fill

        ws.freeze_panes = "A9"

        ws.auto_filter.ref = (
            f"A8:{get_column_letter(len(self.get_headers()))}"
            f"{start + len(rows) - 1}"
        )

    def _format_columns(self, ws):

        for column_cells in ws.columns:

            length = 0

            letter = get_column_letter(
                column_cells[0].column
            )

            for cell in column_cells:

                if cell.value:

                    length = max(
                        length,
                        len(str(cell.value))
                    )

            ws.column_dimensions[letter].width = max(
                length + 3,
                18,
            )

    def _write_footer(self, ws):

        row = ws.max_row + 2

        cell = ws.cell(
            row=row,
            column=1,
        )

        cell.value = self.get_footer()

        cell.font = Font(
            bold=True,
            color="FFFFFF",
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="F97316",
        )

        cell.alignment = Alignment(
            horizontal="center",
        )

        total_columns = len(self.get_headers())

        ws.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=total_columns,
        )
        ws.cell(
            row=row,
            column=1,
        ).alignment = Alignment(
            horizontal="center",
            vertical="center",
        )